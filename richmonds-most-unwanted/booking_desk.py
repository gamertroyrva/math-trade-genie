#!/usr/bin/env python3
"""
booking_desk.py — Richmond's Most Unwanted

Takes the single-tab alphabetical Rap Sheet produced by rap_sheet_builder.py
and books it into the three-tab shareable workbook:

    Tab 1: Raw Data       — the input, untouched
    Tab 2: Hall of Shame  — never-wanted titles only, strikes desc, then alpha
    Tab 3: Full Records   — all titles with Strikes and Wants columns,
                            strikes desc, wants asc, then alpha

On Tabs 2 and 3, NOT_LISTED cells are rendered as blanks for readability.
Tab 1 keeps them — raw data is raw data.

The input file is never modified. All derived numbers are written as values,
never formulas. Trade columns are detected dynamically: every column after
"Title" is treated as a trade. When trade seven's column appears in the
Builder's output, this script needs no changes.

Usage:
    python booking_desk.py rap_sheet.xlsx
    python booking_desk.py rap_sheet.xlsx --out my_output_name.xlsx

Dependency: openpyxl
"""

import argparse
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

STATUS_WANTED = "WANTED"
STATUS_UNWANTED = "UNWANTED"
STATUS_NOT_LISTED = "NOT_LISTED"
VALID_STATUSES = {STATUS_WANTED, STATUS_UNWANTED, STATUS_NOT_LISTED}

TAB_RAW = "Raw Data"
TAB_HALL = "Hall of Shame"
TAB_FULL = "Full Records"

DEFAULT_OUTPUT_NAME = "richmonds-most-unwanted-rap-sheet.xlsx"

# Style guide (Troy's spreadsheet-style-preferences.md, encoded once, here)
FONT_NAME = "Arial"
FONT_SIZE = 10
HEADER_FILL_COLOR = "BFBFBF"  # mid-tone gray
GRID_COLOR = "B0B0B0"

COL_WIDTHS = {"Title": 52, "Strikes": 9, "Wants": 9}
DEFAULT_COL_WIDTH = 13  # trade columns

HEADER_FONT = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color="000000")
DATA_FONT = Font(name=FONT_NAME, size=FONT_SIZE)
HEADER_FILL = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
THIN = Side(style="thin", color=GRID_COLOR)
GRID_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Columns whose content is center-justified (numeral count columns).
# Everything else is left-justified per the style guide.
CENTERED_COLUMNS = {"Strikes", "Wants"}


# ---------------------------------------------------------------------------
# Loading and validation
# ---------------------------------------------------------------------------

def load_rap_sheet(path: Path):
    """Read the Builder's single-tab workbook.

    Returns (trade_columns, rows) where rows is a list of dicts:
    {"Title": str, "statuses": {trade: status}}.

    Hard-stops on any structural surprise — no guessing, no backfilling.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    if len(wb.sheetnames) != 1:
        fail(f"Expected exactly one sheet in {path.name}, "
             f"found {len(wb.sheetnames)}: {wb.sheetnames}")
    ws = wb[wb.sheetnames[0]]

    header = [c.value for c in ws[1]]
    if not header or header[0] != "Title":
        fail(f"Expected first header cell to be 'Title', found: {header[:1]}")

    trade_columns = [str(h) for h in header[1:] if h is not None]
    if len(trade_columns) < 1:
        fail("No trade columns found after 'Title'.")
    if len(set(trade_columns)) != len(trade_columns):
        fail(f"Duplicate trade column names in header: {trade_columns}")

    rows = []
    problems = []
    seen_titles = {}
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True),
                                start=2):
        # Skip fully empty rows (e.g., a trailing blank row)
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        title = row[0]
        if title is None or str(title).strip() == "":
            problems.append(f"Row {r_idx}: blank Title with non-empty row.")
            continue
        title = str(title)
        if title in seen_titles:
            problems.append(
                f"Row {r_idx}: duplicate title '{title}' "
                f"(first seen row {seen_titles[title]}).")
        else:
            seen_titles[title] = r_idx

        statuses = {}
        for j, trade in enumerate(trade_columns, start=1):
            value = row[j] if j < len(row) else None
            value = str(value).strip() if value is not None else ""
            if value not in VALID_STATUSES:
                problems.append(
                    f"Row {r_idx}, column '{trade}': "
                    f"invalid status '{value}' for '{title}'.")
                value = None
            statuses[trade] = value
        rows.append({"Title": title, "statuses": statuses})

    if problems:
        fail("Input validation failed:\n  " + "\n  ".join(problems))
    if not rows:
        fail("No data rows found in input.")

    return trade_columns, rows


# ---------------------------------------------------------------------------
# Derivations (values, never formulas)
# ---------------------------------------------------------------------------

def derive_counts(rows, trade_columns):
    """Attach Strikes, Wants, and Appearances to each row dict."""
    for row in rows:
        statuses = [row["statuses"][t] for t in trade_columns]
        row["Strikes"] = sum(1 for s in statuses if s == STATUS_UNWANTED)
        row["Wants"] = sum(1 for s in statuses if s == STATUS_WANTED)
        row["Appearances"] = sum(
            1 for s in statuses if s != STATUS_NOT_LISTED)


# ---------------------------------------------------------------------------
# Self-QA — loud stops, never shrugs
# ---------------------------------------------------------------------------

def run_qa(rows, hall_rows, full_rows):
    checks = []

    # Every rap-sheet title must have at least one strike, by definition.
    zero_strike = [r["Title"] for r in rows if r["Strikes"] == 0]
    checks.append(("Every title has >= 1 strike",
                   not zero_strike,
                   f"Titles with zero strikes: {zero_strike[:5]}"))

    # Strikes + Wants must equal appearances for every row.
    mismatched = [r["Title"] for r in rows
                  if r["Strikes"] + r["Wants"] != r["Appearances"]]
    checks.append(("Strikes + Wants == Appearances for every title",
                   not mismatched,
                   f"Mismatched titles: {mismatched[:5]}"))

    # Hall of Shame must contain zero wanted appearances.
    tainted = [r["Title"] for r in hall_rows if r["Wants"] != 0]
    checks.append(("Hall of Shame contains only never-wanted titles",
                   not tainted,
                   f"Tainted titles: {tainted[:5]}"))

    # Row counts must reconcile.
    checks.append(("Full Records row count equals Raw Data row count",
                   len(full_rows) == len(rows),
                   f"Full={len(full_rows)}, Raw={len(rows)}"))
    hall_expected = sum(1 for r in rows if r["Wants"] == 0)
    checks.append(("Hall of Shame row count matches never-wanted count",
                   len(hall_rows) == hall_expected,
                   f"Hall={len(hall_rows)}, expected={hall_expected}"))

    failures = [(name, detail) for name, passed, detail in checks
                if not passed]
    if failures:
        lines = [f"  FAILED: {name} — {detail}" for name, detail in failures]
        fail("Self-QA failed. Nothing was written.\n" + "\n".join(lines))

    return [name for name, _, _ in checks]


# ---------------------------------------------------------------------------
# Writing
# ---------------------------------------------------------------------------

def write_tab(wb, name, column_names, data_rows, blank_not_listed):
    """Write one formatted tab. data_rows is a list of lists of cell values."""
    ws = wb.create_sheet(name)

    for j, col in enumerate(column_names, start=1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = GRID_BORDER
        halign = "center" if col in CENTERED_COLUMNS else "left"
        cell.alignment = Alignment(horizontal=halign, vertical="center")

    for i, values in enumerate(data_rows, start=2):
        for j, (col, value) in enumerate(zip(column_names, values), start=1):
            if blank_not_listed and value == STATUS_NOT_LISTED:
                value = None
            cell = ws.cell(row=i, column=j, value=value)
            cell.font = DATA_FONT
            cell.border = GRID_BORDER
            halign = "center" if col in CENTERED_COLUMNS else "left"
            cell.alignment = Alignment(horizontal=halign, vertical="center")

    # Trailing genuinely empty, gridlined row — a visual "more rows welcome"
    trailing = len(data_rows) + 2
    for j in range(1, len(column_names) + 1):
        ws.cell(row=trailing, column=j).border = GRID_BORDER

    ws.freeze_panes = "A2"  # frozen header row; no column freezing
    for j, col in enumerate(column_names, start=1):
        letter = get_column_letter(j)
        ws.column_dimensions[letter].width = COL_WIDTHS.get(
            col, DEFAULT_COL_WIDTH)


def build_workbook(trade_columns, rows, out_path: Path):
    # Tab 1: Raw Data — input order, input content, untouched
    raw_cols = ["Title"] + trade_columns
    raw_data = [[r["Title"]] + [r["statuses"][t] for t in trade_columns]
                for r in rows]

    # Tab 2: Hall of Shame — never wanted, strikes desc, then alpha
    hall_rows = sorted(
        (r for r in rows if r["Wants"] == 0),
        key=lambda r: (-r["Strikes"], r["Title"].lower()))
    hall_cols = ["Title", "Strikes"] + trade_columns
    hall_data = [[r["Title"], r["Strikes"]]
                 + [r["statuses"][t] for t in trade_columns]
                 for r in hall_rows]

    # Tab 3: Full Records — all titles, strikes desc, wants asc, then alpha
    full_rows = sorted(
        rows,
        key=lambda r: (-r["Strikes"], r["Wants"], r["Title"].lower()))
    full_cols = ["Title", "Strikes", "Wants"] + trade_columns
    full_data = [[r["Title"], r["Strikes"], r["Wants"]]
                 + [r["statuses"][t] for t in trade_columns]
                 for r in full_rows]

    qa_passed = run_qa(rows, hall_rows, full_rows)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_tab(wb, TAB_RAW, raw_cols, raw_data, blank_not_listed=False)
    write_tab(wb, TAB_HALL, hall_cols, hall_data, blank_not_listed=True)
    write_tab(wb, TAB_FULL, full_cols, full_data, blank_not_listed=True)
    wb.save(out_path)

    return {
        "raw": len(raw_data),
        "hall": len(hall_data),
        "full": len(full_data),
        "trades": trade_columns,
        "qa": qa_passed,
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def fail(message):
    print(f"booking_desk: HARD STOP\n{message}", file=sys.stderr)
    sys.exit(1)


def main():
    parser = argparse.ArgumentParser(
        description="Book the Rap Sheet into its three-tab "
                    "shareable workbook.")
    parser.add_argument("input", help="Path to the Builder's single-tab "
                                      "rap sheet .xlsx")
    parser.add_argument("--out", default=None,
                        help=f"Output path (default: {DEFAULT_OUTPUT_NAME} "
                             f"beside the input)")
    args = parser.parse_args()

    in_path = Path(args.input)
    if not in_path.exists():
        fail(f"Input file not found: {in_path}")
    out_path = (Path(args.out) if args.out
                else in_path.parent / DEFAULT_OUTPUT_NAME)
    if out_path.resolve() == in_path.resolve():
        fail("Output path equals input path. The input is never modified.")

    trade_columns, rows = load_rap_sheet(in_path)
    derive_counts(rows, trade_columns)
    summary = build_workbook(trade_columns, rows, out_path)

    print(f"booking_desk: booked {summary['raw']} titles "
          f"across {len(summary['trades'])} trades "
          f"({summary['trades'][0]} .. {summary['trades'][-1]})")
    print(f"  {TAB_RAW}: {summary['raw']} rows")
    print(f"  {TAB_HALL}: {summary['hall']} rows")
    print(f"  {TAB_FULL}: {summary['full']} rows")
    print(f"  Self-QA passed: {len(summary['qa'])} checks")
    print(f"  Output: {out_path}")


if __name__ == "__main__":
    main()
