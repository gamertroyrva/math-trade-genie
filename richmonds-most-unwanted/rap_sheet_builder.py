#!/usr/bin/env python3
"""
Richmond's Most Unwanted — Rap Sheet Builder

Sibling tool to Math Trade Genie's numbered pipeline (00-03), NOT part of it.
Input is a Most Wanted/Unwanted GeekList export (a pre-completion demand
snapshot), not an OLWLG results/wants file, so it shares no data contract
with Blocks 0-3. Run this once per math trade, after the final run.

Ingests every `most_wanted_*.txt` file in a folder — one per RVA No-Ship
math trade — and produces:
  1. Per-trade QA metrics, printed to the console.
  2. The Rap Sheet: an .xlsx tracking, for every title that was UNWANTED
     at least once, its status (WANTED / UNWANTED / NOT_LISTED) in every
     trade.

Always a full rebuild from the complete set of raw files — no incremental
append, no diffing against a previous run.

Usage:
  python rap_sheet_builder.py [folder]

If folder is omitted, you'll be prompted for it. Trade IDs are inferred
from filenames: most_wanted_2023_03.txt -> trade ID "2023_03". Trades are
ordered chronologically by sorting these IDs as strings, which works
because they're zero-padded YYYY_MM.
"""

import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


# ── Parsing ───────────────────────────────────────────────────────────────────

FILENAME_PREFIX = "most_wanted_"

MOST_WANTED_HEADER = "MOST WANTED"
UNWANTED_HEADER_RE = re.compile(r"^UNWANTED\s*\((\d+)\)\s*$")

# When one or more traders never submitted a want list, OLWLG appends a
# "MISSING WANT LISTS BY USER (N)" + "MISSING WANT LISTS (N)" cross-reference
# between the ranked list and UNWANTED. It's a redundant re-index of items
# already represented by their "***" line up in the ranked list (no title/
# demand information of its own), so it's skipped entirely rather than parsed.
MISSING_WANT_LISTS_HEADER_RE = re.compile(r"^MISSING WANT LISTS BY USER\s*\(\d+\)\s*$")

# Leading field is either a numbered rank ("10/254/2") or "***" (no want
# list found for the item) followed by whatever want-count numbers remain
# ("*** 5/133/"). Either way we don't need those numbers — just consume
# them generically up to the " - " that precedes title/(id)/[owner].
LISTING_PREFIX = r"[\d/*\s]+-\s*"

MW_PATTERN = re.compile(
    LISTING_PREFIX + r"(?P<title>.*?)\s*\(\s*(?P<id>\d+)\s*\)\s*\[\s*(?P<owner>[^\]]*)\s*\]\s*$"
)
UW_PATTERN = re.compile(
    r"(?P<title>.*?)\s*\(\s*(?P<id>\d+)\s*\)\s*\[\s*(?P<owner>[^\]]*)\s*\]\s*-\s*\d*\s*$"
)

ALT_NAME_PREFIX = "Alt Name:"


class TradeResult:
    def __init__(self, trade_id):
        self.trade_id = trade_id
        self.geeklist_items = 0
        self.wanted_titles = set()
        self.unwanted_titles_raw = set()
        self.malformed_lines = []       # (section, line_no, raw_text) — parsed but title-less
        self.unparseable_lines = []     # (section, line_no, raw_text) — didn't match grammar at all
        self.declared_unwanted_count = None
        self.actual_unwanted_line_count = None

    @property
    def unwanted_titles(self):
        return self.unwanted_titles_raw - self.wanted_titles

    @property
    def overridden_titles(self):
        return self.wanted_titles & self.unwanted_titles_raw

    @property
    def total_unique_titles(self):
        return len(self.wanted_titles) + len(self.unwanted_titles)


def _process_section(lines, pattern, section_name, result, title_sink):
    """Parse one section's lines, updating result and adding valid titles to title_sink."""
    for line_no, raw_line in lines:
        line = raw_line.strip()
        if not line:
            continue

        match = pattern.match(line)
        if match is None:
            result.unparseable_lines.append((section_name, line_no, raw_line))
            continue

        result.geeklist_items += 1
        title = match.group("title").strip()

        if not title:
            result.malformed_lines.append((section_name, line_no, raw_line))
            continue

        if title.startswith(ALT_NAME_PREFIX):
            continue

        title_sink.add(title)


def parse_trade_file(path: Path) -> TradeResult:
    trade_id = path.stem
    if trade_id.startswith(FILENAME_PREFIX):
        trade_id = trade_id[len(FILENAME_PREFIX):]

    result = TradeResult(trade_id)
    text = path.read_text(encoding="utf-8")
    lines = text.splitlines()

    mw_idx = next((i for i, l in enumerate(lines) if l.strip() == MOST_WANTED_HEADER), None)
    if mw_idx is None:
        raise ValueError(f"{path.name}: could not find '{MOST_WANTED_HEADER}' header")

    uw_idx = None
    declared_count = None
    for i in range(mw_idx + 1, len(lines)):
        m = UNWANTED_HEADER_RE.match(lines[i].strip())
        if m:
            uw_idx = i
            declared_count = int(m.group(1))
            break
    if uw_idx is None:
        raise ValueError(f"{path.name}: could not find 'UNWANTED (N)' header")

    mw_end = uw_idx
    for i in range(mw_idx + 1, uw_idx):
        if MISSING_WANT_LISTS_HEADER_RE.match(lines[i].strip()):
            mw_end = i
            break

    mw_lines = [(i + 1, lines[i]) for i in range(mw_idx + 1, mw_end)]
    uw_lines = [(i + 1, lines[i]) for i in range(uw_idx + 1, len(lines))]

    result.declared_unwanted_count = declared_count
    result.actual_unwanted_line_count = sum(1 for _, l in uw_lines if l.strip())

    _process_section(mw_lines, MW_PATTERN, "MOST WANTED", result, result.wanted_titles)
    _process_section(uw_lines, UW_PATTERN, "UNWANTED", result, result.unwanted_titles_raw)

    return result


# ── QA reporting ──────────────────────────────────────────────────────────────

def print_qa_report(results):
    print("=" * 78)
    print("PER-TRADE QA METRICS")
    print("=" * 78)

    for r in results:
        print()
        print(f"[{r.trade_id}]")
        print("-" * 78)
        print(f"  GeekList Items in the trade:   {r.geeklist_items}")
        print(f"  Total unique game titles:      {r.total_unique_titles}")
        print(f"  Wanted:                        {len(r.wanted_titles)}")
        print(f"  Unwanted:                      {len(r.unwanted_titles)}")

        if len(r.wanted_titles) + len(r.unwanted_titles) != r.total_unique_titles:
            print("  *** SANITY CHECK FAILED: Wanted + Unwanted != Total unique titles ***")

        if r.declared_unwanted_count != r.actual_unwanted_line_count:
            print(
                f"  *** WARNING: UNWANTED header declared {r.declared_unwanted_count} lines, "
                f"but {r.actual_unwanted_line_count} were found. Possible truncated/corrupted file. ***"
            )

        if r.overridden_titles:
            print(f"  Rule 1 overrides (appeared in both sections, counted WANTED): {len(r.overridden_titles)}")
            for title in sorted(r.overridden_titles):
                print(f"    - {title}")

        if r.malformed_lines:
            print(f"  Malformed lines (parsed but no title — counted in GeekList Items, excluded from title counts): {len(r.malformed_lines)}")
            for section, line_no, raw_text in r.malformed_lines:
                print(f"    - {section} line {line_no}: {raw_text!r}")

        if r.unparseable_lines:
            print(f"  Unparseable lines (did not match listing grammar at all — excluded from GeekList Items): {len(r.unparseable_lines)}")
            for section, line_no, raw_text in r.unparseable_lines:
                print(f"    - {section} line {line_no}: {raw_text!r}")


def print_near_duplicate_report(results):
    groups = {}
    for r in results:
        for title in (r.wanted_titles | r.unwanted_titles_raw):
            key = re.sub(r"\s+", " ", title).strip().lower()
            groups.setdefault(key, set()).add(title)

    dupes = {k: v for k, v in groups.items() if len(v) > 1}
    if not dupes:
        return

    print()
    print("=" * 78)
    print("NEAR-DUPLICATE TITLES (differ only by case/whitespace — not auto-merged)")
    print("=" * 78)
    for key, variants in sorted(dupes.items()):
        print(f"  - {sorted(variants)}")


# ── Rap Sheet ─────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="000000")
BODY_FONT = Font(name="Arial")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")


def build_rap_sheet(results, out_path: Path):
    analytical_universe = set()
    for r in results:
        analytical_universe |= r.unwanted_titles

    wb = Workbook()
    ws = wb.active
    ws.title = "Rap Sheet"

    trade_ids = [r.trade_id for r in results]
    headers = ["Title"] + trade_ids
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = LEFT_ALIGN

    for title in sorted(analytical_universe, key=str.casefold):
        row = [title]
        for r in results:
            if title in r.wanted_titles:
                row.append("WANTED")
            elif title in r.unwanted_titles:
                row.append("UNWANTED")
            else:
                row.append("NOT_LISTED")
        ws.append(row)

    # One trailing genuinely-empty gridlined row. ws.append([]) is a no-op in
    # openpyxl (no cells get touched, so the row never materializes) — touch
    # each cell explicitly instead so the row exists with no value.
    trailing_row = ws.max_row + 1
    for col in range(1, len(headers) + 1):
        ws.cell(row=trailing_row, column=col)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = LEFT_ALIGN

    title_width = max([len("Title")] + [len(t) for t in analytical_universe]) + 2
    ws.column_dimensions["A"].width = min(title_width, 80)
    for idx, trade_id in enumerate(trade_ids, start=2):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = max(len(trade_id), len("NOT_LISTED")) + 2

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = True

    wb.save(out_path)
    return len(analytical_universe)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except AttributeError:
        pass

    if len(sys.argv) >= 2:
        folder = Path(sys.argv[1])
    else:
        folder = Path(input("Folder containing the most_wanted_*.txt files: ").strip().strip('"'))

    if not folder.is_dir():
        print(f"Error: not a folder: {folder}")
        sys.exit(1)

    files = sorted(folder.glob(f"{FILENAME_PREFIX}*.txt"), key=lambda p: p.stem)
    if not files:
        print(f"Error: no '{FILENAME_PREFIX}*.txt' files found in {folder}")
        sys.exit(1)

    print(f"Found {len(files)} trade file(s):")
    for f in files:
        print(f"  - {f.name}")

    results = [parse_trade_file(f) for f in files]

    print_qa_report(results)
    print_near_duplicate_report(results)

    out_path = folder / "rap_sheet.xlsx"
    universe_size = build_rap_sheet(results, out_path)

    print()
    print("=" * 78)
    print(f"Rap Sheet written: {out_path}")
    print(f"  {universe_size} titles in the Analytical Universe (UNWANTED at least once).")


if __name__ == "__main__":
    main()
